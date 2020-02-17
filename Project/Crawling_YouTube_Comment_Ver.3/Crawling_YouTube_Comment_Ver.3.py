import time
import openpyxl



from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common import exceptions

print("https://github.com/kkm8257    by 김경민")

url=input('댓글을 수집할 사이트 주소 입력 : ')
tempCount=input('해당 주소의 댓글 갯수 : ')
tempCount=int(tempCount)

driver = webdriver.Chrome()
driver.implicitly_wait(3)
driver.get(url)
#스크롤 한번에 20개 정도의 댓글이 있다고 생각, 처음 3번은 댓글창까지 스크롤을 내리는 데 필요한 횟수, 댓글 20개가 로딩되고
#스크롤 세번내려야 댓글 20개가 다 보인후 새로운 댓글 리롤됨


#실제로 유투브 댓글 수가 정확하지 않을 수 있다고 나와있기 때문에, 정확하다고 가정,  댓글 갯수를 표시하는 란은 댓글에 달리는 댓글까지 세는 걸로 예상됨
# https://support.google.com/youtube/answer/7281348?hl=ko   해당링크

def commentCount(tempCount):
    Count=0
    if (tempCount%20 == 0) :
        Count=(tempCount//20)*3 + 3 
        #뒤에 3은 댓글창까지 내려가는데 필요한 대략의 스크롤수
        
        return Count

    else :
        Count=(tempCount//20)*3 + 6
        #뒤에 4는 댓글창까지 내려가는데 필요한 대략의 스크롤 수 + 20으로 나누어떨어지지않았을 때 남은 댓글까지 스크롤하기위한 세번의 3번의 스크롤
       
        return Count


scroll=commentCount(tempCount)

all= driver.find_element_by_tag_name('body')

time.sleep(2)

while scroll:
    
    all.send_keys(Keys.PAGE_DOWN)
    time.sleep(3) #댓글 리롤 대기시간
    scroll-=1
    
  
Id_list=[]
Contents_list=[]


id_contents=driver.find_elements_by_xpath('//*[@id="author-text"]/span')

Contents= driver.find_elements_by_id('content-text')


#id 담기
for temp in id_contents:
    Id_list.append(temp.text)

#댓글 내용 담기
for temp in Contents:
    Contents_list.append(temp.text)


for number in range(len(Contents_list)):
    print(str(number+1) , ".   [ " + Id_list[number] + " ] ,  작성내용 -> [ " + Contents_list[number]+ " ]")
    print()
    print()


print("댓글 갯수 : " + str(len(Contents_list)))
print()
print()



#댓글 검색을 위한 리스트 생성

##########################################################################

Range=len(Contents_list)*2

#Cd리스트를 0으로 초기화 함과 동시에 길이는 Range 만큼 한다.  Cd 리스트에 [아이디 1 , 내용 , 아이디 2 , 내용 ,,,,] 이런식으로 담기위함
Cd=[0 for i in range(Range)]

for i in range(0,len(Id_list)):
    a=i*2
    Cd[a]=Id_list[i]

for i in range(1,len(Contents_list)+1):
    a=i*2-1
    Cd[a]=Contents_list[i-1]



#홀수번째에 있는 댓글내용만 담은 Odd_List
for i in range(len(Cd)):
    Odd_List=Cd[1::2]

##########################################################################




while True:


    InputComment = input("찾고자하는 댓글(종료를 원할경우   ***종료***  라고 입력하세요.) : ")
    print()
    print()
    
    if InputComment=="***종료***":
        break

    Find_Number=[]

    for i in Odd_List:
        
        temp=str(i)

        if temp.find(InputComment)>=0:
            
            Save_Comment_Index=Cd.index(temp)

            Find_Number.append(str(Save_Comment_Index))

    # print(Odd_List)

    ##########################################################################
    #찾은 결과를 Result_Comment 리스트에 [아이디 1 , 내용 , 아이디 2 , 내용 ,,,,] 이렇게 담는다.
    Result_Comment=[]

    for i in Find_Number:
        
        A=int(i)
        A_1=int((A-1)/2)
        Result_Comment.append(Id_list[A_1]) ##

        B=str(Contents_list[A_1])
        B=B.replace('\n','')
        
        Result_Comment.append(B)

    ################################################################################
    #결과 도출
    Result_Number=1

    for number in range(len(Result_Comment)):

        

        if number%2==0:
            print(str(Result_Number) , ".   [ " + Result_Comment[number] + " ] ,  작성내용 -> [ " + Result_Comment[number+1]+ " ]")
            print()
            print()
            Result_Number+=1

        
###########엑셀저장##############
try:
    if not(not Result_Comment):

        SaveOrNot=input('찾은 댓글들을 저장하시겠습니까? 저장을 원하시면 y, 그대로 종료하시려면 n'+'\n')

        wb=openpyxl.Workbook()
        sheet=wb.active
        sheet.title='Sample'

        if SaveOrNot=='y':
            for i in range(0,len(Result_Comment),2):
                sheet.cell(row=i+1 , column=1).value=Result_Comment[i]
            
            for i in range(1,len(Result_Comment),2):
                sheet.cell(row=i, column=3).value=Result_Comment[i]

            wb.save(filename='sample.xlsx')

except :
        
        SaveOrNot=input('출력한 모든 댓글들을 저장하시겠습니까? 저장을 원하시면 y, 그대로 종료하시려면 n'+'\n')

        wb=openpyxl.Workbook()
        sheet=wb.active
        sheet.title='Sample'

        if SaveOrNot=='y':
            for i in range(0,len(Cd),2):
                sheet.cell(row=i+1 , column=1).value=Cd[i]
            
            for i in range(1,len(Cd),2):
                sheet.cell(row=i, column=3).value=Cd[i]

            wb.save(filename='sample.xlsx')

    
